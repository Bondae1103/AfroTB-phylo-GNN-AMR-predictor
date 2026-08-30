"""Convert 0-StartHERE_Afro-TB.xlsx into a reproducible mutation feature matrix.

Reads the Afro-TB mutation workbook and produces:
  data/processed/features.csv           Name + 157 binary mutation columns
  data/processed/labels.csv             Name, Country, Lineage, Drug (unmodified)
  data/processed/dataset_metadata.json  Dataset-level metadata

PROVISIONAL PREPROCESSING ASSUMPTION
-------------------------------------
The source publication (Laamarti et al., Scientific Data, 2023) and the
workbook itself do not explicitly distinguish the meaning of the three
placeholder values found in the 157 mutation columns: "_", "-", and a
blank/None cell. Per the statistical audit, no column mixes a drug code
with more than one placeholder type inconsistently, and drug codes never
co-occur with a placeholder in the same cell. Lacking an authoritative
definition, this script treats all three placeholder values as equivalent
to "mutation not present" (0), and any recognized drug-resistance code as
"mutation present" (1). This is a simplification that should be revisited
if authoritative documentation (e.g. the UM6P Afro-TB database) becomes
available, since "_" / "-" / blank could in principle distinguish
"not detected" from "not tested" or "not applicable".

This script does not modify the raw Excel file, does not read VCFs, does
not normalize Lineage/Drug labels, does not drop any rows, and does not
perform train/test splitting or graph construction.
"""

import json
from pathlib import Path

import openpyxl

RAW_XLSX = Path("data/raw/Afro_TB/0-StartHERE_Afro-TB.xlsx")
OUT_DIR = Path("data/processed")
SHEET_NAME = "AfroTB"

MUTATION_HEADER_ROW = 4  # 157 mutation names, columns 5-161
BASE_HEADER_ROW = 5      # Name, Country, Lineage, Drug, columns 1-4
DATA_START_ROW = 6

MUTATION_COL_START = 5   # 1-based column index (inclusive)
MUTATION_COL_END = 161   # 1-based column index (inclusive)

EXPECTED_N_ISOLATES = 13753
EXPECTED_N_MUTATION_COLS = 157

RECOGNIZED_PRESENCE_CODES = {"RIF", "INH", "EMB", "PZA", "STM", "LEV", "CAP", "ETH", "LZD"}
PLACEHOLDER_CODES = {"_", "-", None}


def load_workbook_rows():
    """Load header rows and data rows from the Afro-TB workbook.

    Returns (mutation_names, base_headers, data_rows) where data_rows is a
    list of raw row tuples starting at DATA_START_ROW, in original order,
    stopping at the first row whose Name cell is None.
    """
    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    mutation_header_row = None
    base_header_row = None
    data_rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == MUTATION_HEADER_ROW:
            mutation_header_row = row
        elif i == BASE_HEADER_ROW:
            base_header_row = row
        elif i >= DATA_START_ROW:
            if row[0] is None:
                continue
            data_rows.append(row)

    wb.close()

    mutation_names = list(mutation_header_row[MUTATION_COL_START - 1:MUTATION_COL_END])
    base_headers = list(base_header_row[0:4])

    return mutation_names, base_headers, data_rows


def validate_structure(mutation_names, base_headers, data_rows):
    """Verify isolate count, mutation-column count, and header identity."""
    if len(data_rows) != EXPECTED_N_ISOLATES:
        raise AssertionError(
            f"Expected {EXPECTED_N_ISOLATES} isolates, found {len(data_rows)}."
        )

    if len(mutation_names) != EXPECTED_N_MUTATION_COLS:
        raise AssertionError(
            f"Expected {EXPECTED_N_MUTATION_COLS} mutation columns, "
            f"found {len(mutation_names)}."
        )

    if any(name is None for name in mutation_names):
        raise AssertionError("One or more mutation column headers (row 4) are blank.")

    expected_base = ["Name", "Country", "Lineage", "Drug"]
    if base_headers != expected_base:
        raise AssertionError(
            f"Row 5 headers {base_headers} do not match expected {expected_base}."
        )


def validate_names_unique(names):
    """Verify that Name values are unique."""
    seen = set()
    duplicates = set()
    for name in names:
        if name in seen:
            duplicates.add(name)
        seen.add(name)
    if duplicates:
        raise AssertionError(f"Duplicate Name values found: {sorted(duplicates)}")


def validate_mutation_values(data_rows, mutation_names):
    """Ensure every mutation cell is either a recognized drug code or a placeholder.

    Raises ValueError naming the offending row/column/value on any other value.
    """
    for row_idx, row in enumerate(data_rows, start=DATA_START_ROW):
        name = row[0]
        mutation_values = row[MUTATION_COL_START - 1:MUTATION_COL_END]
        for col_offset, value in enumerate(mutation_values):
            if value in PLACEHOLDER_CODES:
                continue
            if value in RECOGNIZED_PRESENCE_CODES:
                continue
            mutation_name = mutation_names[col_offset]
            raise ValueError(
                f"Unexpected mutation value {value!r} at spreadsheet row {row_idx} "
                f"(Name={name!r}), mutation column {mutation_name!r}. "
                f"Expected one of {sorted(RECOGNIZED_PRESENCE_CODES)} or a "
                f"placeholder {sorted(str(p) for p in PLACEHOLDER_CODES)}."
            )


def build_features_and_labels(data_rows, mutation_names):
    """Binarize mutation columns and split off the label columns.

    Returns (feature_rows, label_rows) where feature_rows is a list of
    [Name, 0/1, 0/1, ...] and label_rows is a list of [Name, Country,
    Lineage, Drug], both in original row order.
    """
    feature_rows = []
    label_rows = []

    for row in data_rows:
        name, country, lineage, drug = row[0], row[1], row[2], row[3]
        mutation_values = row[MUTATION_COL_START - 1:MUTATION_COL_END]

        binary_values = [
            1 if value in RECOGNIZED_PRESENCE_CODES else 0
            for value in mutation_values
        ]

        feature_rows.append([name] + binary_values)
        label_rows.append([name, country, lineage, drug])

    return feature_rows, label_rows


def write_csv(path, header, rows):
    import csv

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def write_metadata(path, mutation_names, n_isolates):
    metadata = {
        "n_isolates": n_isolates,
        "n_mutation_features": len(mutation_names),
        "feature_names": mutation_names,
        "recognized_presence_codes": sorted(RECOGNIZED_PRESENCE_CODES),
        "placeholder_codes": ["_", "-", None],
        "preprocessing_assumption": (
            "The placeholder values '_', '-', and blank/None in the 157 "
            "mutation columns are treated as equivalent 'mutation absent' "
            "(0) states. This is a provisional assumption: the source "
            "publication and workbook do not explicitly define whether "
            "these three placeholders distinguish 'not detected' from "
            "'not tested' or 'not applicable'. Any recognized drug-"
            "resistance code (RIF, INH, EMB, PZA, STM, LEV, CAP, ETH, LZD) "
            "is treated as 'mutation present' (1)."
        ),
        "source_file": str(RAW_XLSX),
        "mutation_header_row": MUTATION_HEADER_ROW,
        "base_header_row": BASE_HEADER_ROW,
        "data_start_row": DATA_START_ROW,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)


def main():
    mutation_names, base_headers, data_rows = load_workbook_rows()

    validate_structure(mutation_names, base_headers, data_rows)
    validate_names_unique([row[0] for row in data_rows])
    validate_mutation_values(data_rows, mutation_names)

    feature_rows, label_rows = build_features_and_labels(data_rows, mutation_names)

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    write_csv(OUT_DIR / "features.csv", ["Name"] + mutation_names, feature_rows)
    write_csv(OUT_DIR / "labels.csv", ["Name", "Country", "Lineage", "Drug"], label_rows)
    write_metadata(OUT_DIR / "dataset_metadata.json", mutation_names, len(data_rows))

    print(f"Wrote {len(data_rows)} isolates x {len(mutation_names)} mutation features.")
    print(f"  {OUT_DIR / 'features.csv'}")
    print(f"  {OUT_DIR / 'labels.csv'}")
    print(f"  {OUT_DIR / 'dataset_metadata.json'}")


if __name__ == "__main__":
    main()
